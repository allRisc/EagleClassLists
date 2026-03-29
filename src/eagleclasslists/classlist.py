####################################################################################################
# EagleClassLists is a tool used to aid in the creation of class lists for schools.
# Copyright (C) 2026, Benjamin Davis
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
####################################################################################################

"""Data models for class lists, teachers, and students.

This module defines the core data structures used throughout EagleClassLists
for representing grade levels, classrooms, teachers, and students with their
attributes.
"""

from __future__ import annotations

import enum
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, BinaryIO, ClassVar

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


class Gender(enum.StrEnum):
    """Enumeration of student gender options."""

    MALE = "Male"
    """Male gender."""

    FEMALE = "Female"
    """Female gender."""


class Academics(enum.StrEnum):
    """Enumeration of academic performance levels."""

    HIGH = "High"
    """High academic performance."""

    MEDIUM = "Medium"
    """Medium academic performance."""

    LOW = "Low"
    """Low academic performance."""


class Behavior(enum.StrEnum):
    """Enumeration of behavior performance levels."""

    HIGH = "High"
    """High behavior rating (good behavior)."""

    MEDIUM = "Medium"
    """Medium behavior rating."""

    LOW = "Low"
    """Low behavior rating (needs improvement)."""


class Cluster(enum.StrEnum):
    """Enumeration of special program clusters."""

    AC = "AC"
    """Academically Challenged cluster."""

    GEM = "GEM"
    """Gifted Education Model cluster."""

    EL = "EL"
    """English Learner cluster."""


@dataclass
class GradeList:
    """Represents a grade level with its classrooms, teachers, and students.

    This class holds all the data for a single grade level, including the
    list of classrooms, all teachers, and all students at that grade.
    """

    STUDENT_HEADER_MAPPING: ClassVar[dict[str, str]] = {
        "first_name": "First Name",
        "last_name": "Last Name",
        "gender": "Gender",
        "academics": "Academics",
        "behavior": "Behavior",
        "cluster": "Cluster",
        "resource": "Resource",
        "speech": "Speech",
    }

    TEACHER_HEADER_MAPPING: ClassVar[dict[str, str]] = {
        "name": "Name",
        "clusters": "Clusters",
    }

    classes: list[Classroom]
    """List of Classroom objects in this grade."""

    teachers: list[Teacher]
    """List of Teacher objects for this grade."""

    students: list[Student]
    """List of Student objects in this grade."""

    def save_to_excel(self, filepath: str | Path | BinaryIO) -> None:
        """Save the grade list to an Excel file with three sheets.

        Creates an Excel file with "Teachers", "Students", and "Classrooms"
        sheets following the Option C structure.

        Args:
            filepath: Path to the Excel file to create, or a file-like object
                to write to.
        """
        wb = Workbook()

        # Remove default sheet and create our three sheets
        if wb.active is not None:
            wb.remove(wb.active)

        self._save_teachers(wb)
        self._save_students(wb)
        self._save_classrooms(wb)

        wb.save(filepath)

    def _save_teachers(self, wb: Workbook) -> None:
        """Save the list of teachers to a 'Teachers' sheet in the provided workbook.

        Args:
            wb: The excel workbook to save to.
        """
        teachers_sheet = wb.create_sheet("Teachers")

        # Write Teachers sheet
        teachers_sheet.append(
            [
                CellRichText(TextBlock(InlineFont(b=True), heading))
                for heading in self.TEACHER_HEADER_MAPPING.values()
            ]
        )
        for teacher in self.teachers:
            teachers_sheet.append([
                _attr_to_save_str(teacher, attr) for attr in self.TEACHER_HEADER_MAPPING.keys()
            ])

    def _save_students(self, wb: Workbook) -> None:
        """Save the list of students to a 'Students' sheet in the provided workbook.

        Args:
            wb: The excel workbook to save to.
        """
        students_sheet = wb.create_sheet("Students")

        # Write Students sheet
        students_sheet.append(
            [
                CellRichText(TextBlock(InlineFont(b=True), heading))
                for heading in self.STUDENT_HEADER_MAPPING.values()
            ]
        )
        for student in self.students:
            students_sheet.append([
                _attr_to_save_str(student, attr) for attr in self.STUDENT_HEADER_MAPPING.keys()
            ])

    def _save_classrooms(self, wb: Workbook) -> None:
        """Save a mapping of teachers to students to the 'Classrooms' sheet of the workbook.

        Args:
            wb: The excel workbook to save to.
        """
        classrooms_sheet = wb.create_sheet("Classrooms")

        # Write Classrooms sheet
        classrooms_sheet.append(["teacher_name", "student_first", "student_last"])
        for classroom in self.classes:
            for student in classroom.students:
                classrooms_sheet.append(
                    [classroom.teacher.name, student.first_name, student.last_name]
                )

    @classmethod
    def from_excel(cls, filepath: str | Path) -> GradeList:
        """Load a grade list from an Excel file.

        Reads an Excel file with "Teachers", "Students", and "Classrooms"
        sheets and reconstructs the GradeList object.

        Args:
            filepath: Path to the Excel file to read.

        Returns:
            A GradeList object populated from the Excel data.

        Raises:
            ValueError: If the Excel file has invalid data.
        """
        wb = load_workbook(filepath)

        teachers = cls._load_teachers_dict(wb)
        students = cls._load_students_dict(wb)
        classrooms = cls._load_classroms(wb, students, teachers)

        # Update teacher references in students
        for classroom in classrooms:
            for student in classroom.students:
                student.teacher = classroom.teacher

        return cls(
            classes=classrooms,
            teachers=list(teachers.values()),
            students=list(students.values()),
        )

    @staticmethod
    def _load_teachers_dict(wb: Workbook) -> dict[str, Teacher]:
        """Load a dictionary with the teacher definitions from the excel workbook.

        Args:
            wb: The excel workbook to load from.
        """
        teachers_sheet = wb["Teachers"]
        teachers: dict[str, Teacher] = {}
        for row in teachers_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            name = str(row[0])
            clusters_str = str(row[1]) if row[1] else ""
            cluster_list = [Cluster(c.strip()) for c in clusters_str.split(",") if c.strip()]
            teachers[name] = Teacher(name=name, clusters=cluster_list)

        return teachers

    @staticmethod
    def _load_students_dict(wb: Workbook) -> dict[tuple[str, str], Student]:
        """Load a dictionary with the student definitions from the excel workbook.

        Args:
            wb: The excel workbook to load from.
        """
        # Read Students sheet
        students_sheet = wb["Students"]
        students: dict[tuple[str, str], Student] = {}
        for row in students_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            first_name = str(row[0])
            last_name = str(row[1])
            gender = Gender(str(row[2])) if row[2] else Gender.MALE
            academics = Academics(str(row[3])) if row[3] else Academics.MEDIUM
            behavior = Behavior(str(row[4])) if row[4] else Behavior.MEDIUM
            cluster = Cluster(str(row[5])) if row[5] else None
            resource = str(row[6]).upper() == "TRUE" if row[6] else False
            speech = str(row[7]).upper() == "TRUE" if row[7] else False

            student = Student(
                first_name=first_name,
                last_name=last_name,
                gender=gender,
                academics=academics,
                behavior=behavior,
                cluster=cluster,
                resource=resource,
                speech=speech,
            )
            students[(first_name, last_name)] = student
        return students

    @staticmethod
    def _load_classroms(
        wb: Workbook,
        students: dict[tuple[str, str], Student],
        teachers: dict[str, Teacher]
    ) -> list[Classroom]:
        """Load a list of classrooms based on the excel file and the previously found teachers and
        students.

        Args:
            wb: The excel workbook to load from.
            students: The dictionary mapping first and last names of students to their corresponding
                Student object.
            teachers: The dictionary mapping of teacher names to the corresponding Teacher object.

        Returns:
            A list of classroom objects

        Raises:
            ValueError: If a specified student or teacher is not valid.
        """
        # Read Classrooms sheet and build classroom structure
        classrooms_sheet = wb["Classrooms"]
        classroom_map: dict[str, list[Student]] = {}
        for row in classrooms_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1] or not row[2]:
                continue
            teacher_name = str(row[0])
            student_first = str(row[1])
            student_last = str(row[2])

            student_key = (student_first, student_last)
            if student_key in students:
                if teacher_name not in classroom_map:
                    classroom_map[teacher_name] = []
                classroom_map[teacher_name].append(students[student_key])

        # Build Classroom objects
        classroom_list: list[Classroom] = []
        for teacher_name, student_list in classroom_map.items():
            if teacher_name in teachers:
                classroom_list.append(
                    Classroom(teacher=teachers[teacher_name], students=student_list)
                )

        return classroom_list


@dataclass
class Classroom:
    """Represents a single classroom with a teacher and assigned students."""

    teacher: Teacher
    """The Teacher assigned to this classroom."""

    students: list[Student]
    """List of Student objects assigned to this classroom."""


@dataclass
class Teacher:
    """Represents a teacher with their name and cluster qualifications."""

    name: str
    """The teacher's full name."""

    clusters: list[Cluster] = field(default_factory=list)
    """List of Cluster types this teacher is qualified to teach."""


@dataclass
class Student:
    """Represents a student with their personal and academic attributes."""

    first_name: str
    """The student's first name."""

    last_name: str
    """The student's last name."""

    gender: Gender
    """The student's gender (Gender enum)."""

    academics: Academics
    """The student's academic performance level (Academics enum)."""

    behavior: Behavior
    """The student's behavior rating (Behavior enum)."""

    teacher: Teacher | None = None
    """The assigned Teacher, or None if unassigned."""

    cluster: Cluster | None = None
    """The student's Cluster assignment, or None."""

    resource: bool = False
    """Whether the student receives resource services."""

    speech: bool = False
    """Whether the student receives speech services."""


def _attr_to_save_str(obj: Any, attr: str) -> str:
    if attr != "" and not hasattr(obj, attr):
        raise ValueError(f"Object of type {type(obj).__name__} does not have attr '{attr}'")

    if attr == "":
        val = obj
    else:
        val = getattr(obj, attr)

    if val is None:
        return ""

    if isinstance(val, str):
        return val

    if isinstance(val, list):
        return ", ".join([_attr_to_save_str(item, "") for item in val])

    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"

    return str(val)
