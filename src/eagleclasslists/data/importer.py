####################################################################################################
# EagleClassLists is a tool used to aid in the creation of class lists for schools.
# Copyright (C) 2026, Benjamin Davis
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
####################################################################################################

"""Excel import/export functions for classlist data.

This module provides standalone functions for reading and writing teachers,
students, and classroom data to/from Excel files, decoupled from the data models.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
import pydantic

from eagleclasslists.data.classlist import Classroom, Student, Teacher
from eagleclasslists.data.errors import ExcelImportError
from eagleclasslists.data.settings import (
    DEFAULT_PRESET,
    ColumnMappingPreset,
    rename_records_for_reading,
    rename_records_for_writing,
)
from eagleclasslists.data.types import _pydantic_bool_parser


def _merge_split_cluster_columns(
    records: list[dict[str, Any]],
    split_map: dict[str, str],
) -> list[dict[str, Any]]:
    """Merge split boolean cluster columns into a single ``cluster`` key.

    Called after :func:`rename_records_for_reading`.  The split column headers
    are *not* in the regular ``student_columns`` mapping, so they pass through
    the rename step unchanged and are still keyed by their Excel header names.

    Args:
        records: Records that have already been through rename (Python attr
            names for standard columns, Excel headers for split columns).
        split_map: Maps Cluster enum value (``"AC"``) to Excel column header.

    Returns:
        The same records list, mutated in-place, with split column keys removed
        and ``"cluster"`` set when exactly one split column is truthy.

    Raises:
        ExcelImportError: If more than one split column is truthy for a row.
    """
    reverse: dict[str, str] = {header: cluster for cluster, header in split_map.items()}

    for record in records:
        matched: list[str] = []
        for header, cluster_val in reverse.items():
            raw = record.pop(header, None)
            if raw is None:
                continue
            try:
                if _pydantic_bool_parser(raw):
                    matched.append(cluster_val)
            except ValueError:
                pass

        if len(matched) > 1:
            first_name = record.get("first_name", "UNKNOWN")
            last_name = record.get("last_name", "UNKNOWN")
            raise ExcelImportError(
                "Multiple cluster columns are set for a student",
                f"Student {first_name} {last_name} has multiple cluster columns marked as Yes: "
                f"{', '.join(matched)}. Only one cluster is allowed per student.",
            )
        if len(matched) == 1:
            record["cluster"] = matched[0]

    return records


def _expand_split_cluster_columns(
    records: list[dict[str, Any]],
    split_map: dict[str, str],
    cluster_header: str | None,
) -> list[dict[str, Any]]:
    """Expand a single cluster column into split boolean columns for writing.

    Called after :func:`rename_records_for_writing` so keys are Excel headers.

    Args:
        records: Records with Excel column header keys.
        split_map: Maps Cluster enum value (``"AC"``) to Excel column header.
        cluster_header: The Excel column header for the original cluster column
            (from ``student_columns["cluster"]``), used to read and optionally
            remove the original value.

    Returns:
        The same records list, mutated in-place, with boolean split columns
        added and the original cluster column removed when the value is
        represented by a split column.
    """
    for record in records:
        cluster_value = record.get(cluster_header) if cluster_header else None
        # Normalise enum instances and strings to a plain string for comparison
        cluster_str = str(cluster_value) if cluster_value is not None else None

        for cluster_enum_val, header in split_map.items():
            record[header] = "Yes" if cluster_str == cluster_enum_val else "No"

        # Remove the original cluster column only when the value is covered
        # by one of the split columns (or is None/empty).
        if cluster_header and cluster_header in record:
            if cluster_str is None or cluster_str in split_map:
                del record[cluster_header]

    return records


def save_teachers_to_excel(
    teachers: list[Teacher],
    filepath: str | Path,
    preset: ColumnMappingPreset = DEFAULT_PRESET,
) -> None:
    """Save teachers to a single-sheet Excel file.

    Args:
        teachers: List of Teacher objects to save.
        filepath: Path to the Excel file to create.
        preset: Column mapping preset defining sheet name and column headers.
    """
    records = [t.model_dump(by_alias=False) for t in teachers]
    renamed = rename_records_for_writing(records, preset.teacher_columns)
    df = pd.DataFrame(renamed) if renamed else pd.DataFrame()
    df.to_excel(filepath, sheet_name=preset.teachers_sheet, index=False)


def save_students_to_excel(
    students: list[Student],
    filepath: str | Path,
    preset: ColumnMappingPreset = DEFAULT_PRESET,
) -> None:
    """Save students to a single-sheet Excel file.

    Args:
        students: List of Student objects to save.
        filepath: Path to the Excel file to create.
        preset: Column mapping preset defining sheet name and column headers.
    """
    records = [s.model_dump(by_alias=False) for s in students]
    renamed = rename_records_for_writing(records, preset.student_columns)
    if preset.split_cluster_columns:
        cluster_header = preset.student_columns.get("cluster")
        _expand_split_cluster_columns(renamed, preset.split_cluster_columns, cluster_header)
    df = pd.DataFrame(renamed) if renamed else pd.DataFrame()
    df.to_excel(filepath, sheet_name=preset.students_sheet, index=False)


def save_classrooms_to_excel(
    classes: list[Classroom],
    filepath: str | Path,
    preset: ColumnMappingPreset = DEFAULT_PRESET,
) -> None:
    """Save classroom assignments to a single-sheet Excel file.

    The file contains only assignment mappings (teacher name,
    student first name, student last name).

    Args:
        classes: List of Classroom objects to save.
        filepath: Path to the Excel file to create.
        preset: Column mapping preset defining sheet name and column headers.
    """
    records: list[dict[str, str | None]] = []
    for classroom in classes:
        for student in classroom.students:
            records.append(
                {
                    "teacher_name": classroom.teacher.name,
                    "teacher_grade": classroom.teacher.grade,
                    "student_first_name": student.first_name,
                    "student_last_name": student.last_name,
                    "student_grade": student.grade,
                }
            )
    renamed = rename_records_for_writing(records, preset.classroom_columns)
    df = pd.DataFrame(renamed) if renamed else pd.DataFrame()
    df.to_excel(filepath, sheet_name=preset.classrooms_sheet, index=False)


def update_student_file_with_teachers(
    input_filepath: str | Path,
    output_filepath: str | Path,
    students: list[Student],
    preset: ColumnMappingPreset = DEFAULT_PRESET,
    classrooms: list[Classroom] | None = None,
) -> None:
    """Update a student Excel file with teacher assignments from a student list.

    Reads students from an input Excel file, matches them by (first_name, last_name)
    against the provided student list, and fills in the teacher column for matching
    students. Writes the updated data to an output Excel file.

    Classroom assignments take priority over ``Student.teacher`` when *classrooms*
    is provided.

    Args:
        input_filepath: Path to the input Excel file to read.
        output_filepath: Path to the output Excel file to write.
        students: List of Student objects containing teacher assignments.
        preset: Column mapping preset defining sheet name and column headers.
        classrooms: Optional list of Classroom objects. When provided, the
            teacher for each student is determined by the classroom the student
            is assigned to, falling back to ``Student.teacher`` for students
            not found in any classroom.

    Raises:
        ExcelImportError: If the Excel file cannot be read, parsed, or written.
    """
    try:
        import openpyxl

        teacher_header = preset.student_columns.get("teacher")
        first_name_header = preset.student_columns.get("first_name")
        last_name_header = preset.student_columns.get("last_name")

        # Build lookup dict: {(first_name.lower(), last_name.lower()): teacher_name}
        # Start with Student.teacher as fallback
        teacher_lookup: dict[tuple[str, str], str] = {}
        for student in students:
            if student.teacher:
                key = (student.first_name.lower(), student.last_name.lower())
                teacher_lookup[key] = student.teacher

        # Override with classroom assignments (authoritative source)
        if classrooms:
            for classroom in classrooms:
                for student in classroom.students:
                    key = (student.first_name.lower(), student.last_name.lower())
                    teacher_lookup[key] = classroom.teacher.name

        # Load workbook preserving formatting
        wb = openpyxl.load_workbook(input_filepath)

        # Find the target sheet
        if preset.students_sheet in wb.sheetnames:
            ws = wb[preset.students_sheet]
        elif wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
        else:
            raise ExcelImportError("Could not find a sheet in the provided excel file")

        # Map header names to column indices from the first row
        header_map: dict[str | None, int] = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            header_map[cell.value] = col_idx

        teacher_col = header_map.get(teacher_header)
        first_name_col = header_map.get(first_name_header)
        last_name_col = header_map.get(last_name_header)

        # Add teacher column if absent
        if teacher_header is not None and teacher_col is None:
            teacher_col = ws.max_column + 1
            ws.cell(row=1, column=teacher_col, value=teacher_header)

        # Update teacher cells in-place (data rows start at row 2)
        if teacher_col is not None:
            for row in range(2, ws.max_row + 1):
                first_name_val = (
                    ws.cell(row=row, column=first_name_col).value
                    if first_name_col
                    else None
                )
                last_name_val = (
                    ws.cell(row=row, column=last_name_col).value
                    if last_name_col
                    else None
                )

                if first_name_val is not None and last_name_val is not None:
                    key = (str(first_name_val).lower(), str(last_name_val).lower())
                    teacher_name = teacher_lookup.get(key, "")
                    ws.cell(row=row, column=teacher_col).value = teacher_name
                else:
                    ws.cell(row=row, column=teacher_col).value = ""

        wb.save(output_filepath)

    except FileNotFoundError as e:
        raise ExcelImportError(
            "Excel file not found",
            "The specified file could not be found. Please check the path.",
        ) from e
    except Exception as e:
        if isinstance(e, ExcelImportError):
            raise
        raise ExcelImportError(
            "Unexpected error updating student file",
            f"An unexpected error occurred: {type(e).__name__}: {e}",
        ) from e


def load_teachers_from_excel(
    filepath: str | Path,
    preset: ColumnMappingPreset = DEFAULT_PRESET,
) -> list[Teacher]:
    """Load teachers from a single-sheet Excel file.

    Args:
        filepath: Path to the Excel file to read.
        preset: Column mapping preset defining sheet name and column headers.

    Returns:
        A list of Teacher objects.

    Raises:
        ExcelImportError: If the Excel file has invalid data or format.
    """
    try:
        with pd.ExcelFile(filepath) as ef:
            records = _sheet_to_clean_records(ef, preset.teachers_sheet)
            renamed = rename_records_for_reading(records, preset.teacher_columns)
            return [Teacher.model_validate(r) for r in renamed]
    except FileNotFoundError as e:
        raise ExcelImportError(
            "Excel file not found",
            "The specified file could not be found. Please check the path.",
        ) from e
    except pd.errors.ParserError as e:
        raise ExcelImportError(
            "Could not parse the Excel file",
            f"The file appears to be corrupted or is not a valid Excel file. Technical error: {e}",
        ) from e
    except pydantic.ValidationError as e:
        errors = []
        for err in e.errors():
            loc = " -> ".join(str(x) for x in err["loc"])
            msg = err["msg"]
            errors.append(f"  - {loc}: {msg}")
        error_details = "\n".join(errors)
        raise ExcelImportError(
            "Teacher data validation failed.",
            f"Errors:\n{error_details}\n\nRequired columns: Name, Clusters (optional).",
        ) from e
    except Exception as e:
        if isinstance(e, ExcelImportError):
            raise
        raise ExcelImportError(
            "Unexpected error loading teachers file",
            f"An unexpected error occurred: {type(e).__name__}: {e}",
        ) from e


def load_students_from_excel(
    filepath: str | Path,
    preset: ColumnMappingPreset = DEFAULT_PRESET,
) -> list[Student]:
    """Load students from a single-sheet Excel file.

    Args:
        filepath: Path to the Excel file to read.
        preset: Column mapping preset defining sheet name and column headers.

    Returns:
        A list of Student objects.

    Raises:
        ExcelImportError: If the Excel file has invalid data or format.
    """
    try:
        with pd.ExcelFile(filepath) as ef:
            records = _sheet_to_clean_records(ef, preset.students_sheet)
            renamed = rename_records_for_reading(records, preset.student_columns)
            if preset.split_cluster_columns:
                _merge_split_cluster_columns(renamed, preset.split_cluster_columns)
            # Validate each student individually to get better error messages
            validation_errors = []
            for _i, r in enumerate(renamed):
                first_name = r.get("first_name", "UNKNOWN")
                last_name = r.get("last_name", "UNKNOWN")
                try:
                    Student.model_validate(r)
                except pydantic.ValidationError as ve:
                    # Collect errors with student name
                    for err in ve.errors():
                        loc = err["loc"][0] if err["loc"] else "unknown"
                        if loc not in ("first_name", "last_name"):
                            validation_errors.append(
                                f"  - {first_name} {last_name}: {loc} {err['msg']}"
                            )
            if validation_errors:
                error_details = "\n".join(validation_errors)
                raise ExcelImportError(
                    "Student data validation failed.",
                    f"Errors:\n{error_details}\n\n"
                    f"Required columns: First Name, Last Name, Gender, Math, ELA, Behavior.\n"
                    f"Optional columns: Cluster, Resource, Speech, Teacher, Exclusions.",
                )
            return [Student.model_validate(r) for r in renamed]
    except pydantic.ValidationError as e:
        # Fallback for any validation errors not caught above
        errors = []
        for err in e.errors():
            loc = " -> ".join(str(x) for x in err["loc"])
            msg = err["msg"]
            errors.append(f"  - {loc}: {msg}")
        error_details = "\n".join(errors)
        raise ExcelImportError(
            "Student data validation failed.",
            f"Errors:\n{error_details}\n\n"
            f"Required columns: First Name, Last Name, Gender, Math, ELA, Behavior.\n"
            f"Optional columns: Cluster, Resource, Speech, Teacher, Exclusions.",
        ) from e
    except FileNotFoundError as e:
        raise ExcelImportError(
            "Excel file not found",
            "The specified file could not be found. Please check the path.",
        ) from e
    except pd.errors.ParserError as e:
        raise ExcelImportError(
            "Could not parse the Excel file",
            f"The file appears to be corrupted or is not a valid Excel file. Technical error: {e}",
        ) from e
    except Exception as e:
        if isinstance(e, ExcelImportError):
            raise
        raise ExcelImportError(
            "Unexpected error loading students file",
            f"An unexpected error occurred: {type(e).__name__}: {e}",
        ) from e


def load_classrooms_from_excel(
    filepath: str | Path,
    teachers: list[Teacher],
    students: list[Student],
    preset: ColumnMappingPreset = DEFAULT_PRESET,
) -> list[Classroom]:
    """Load classroom assignments from a single-sheet Excel file.

    Requires teachers and students lists to resolve references.

    Args:
        filepath: Path to the Excel file to read.
        teachers: List of Teacher objects for name resolution.
        students: List of Student objects for name resolution.
        preset: Column mapping preset defining sheet name and column headers.

    Returns:
        A list of Classroom objects.

    Raises:
        ExcelImportError: If references cannot be resolved.
    """
    teacher_dict: dict[str, Teacher] = {t.name: t for t in teachers}
    student_dict: dict[str, Student] = {f"{s.first_name}_{s.last_name}": s for s in students}

    try:
        with pd.ExcelFile(filepath) as ef:
            raw_records = _sheet_to_clean_records(ef, preset.classrooms_sheet)
            records = rename_records_for_reading(raw_records, preset.classroom_columns)
    except FileNotFoundError as e:
        raise ExcelImportError(
            "Excel file not found",
            "The specified file could not be found. Please check the path.",
        ) from e
    except pd.errors.ParserError as e:
        raise ExcelImportError(
            "Could not parse the Excel file",
            f"The file appears to be corrupted or is not a valid Excel file. Technical error: {e}",
        ) from e
    except Exception as e:
        if isinstance(e, ExcelImportError):
            raise
        raise ExcelImportError(
            "Unexpected error loading classrooms file",
            f"An unexpected error occurred: {type(e).__name__}: {e}",
        ) from e

    classrooms: dict[str, Classroom] = {}
    for rec in records:
        teacher_name = rec.get("teacher_name")
        first_name = rec.get("student_first_name")
        last_name = rec.get("student_last_name")

        if not teacher_name:
            raise ExcelImportError(
                "Missing Teacher Name in classroom data",
                "Each row must have a 'teacher_name' column.",
            )
        if not first_name or not last_name:
            raise ExcelImportError(
                "Missing student name in classroom data",
                "Each row must have 'student_first_name' and 'student_last_name' columns.",
            )

        if teacher_name not in teacher_dict:
            raise ExcelImportError(
                f"Teacher '{teacher_name}' not found",
                "Load the teachers file before loading classrooms.",
            )

        student_key = f"{first_name}_{last_name}"
        if student_key not in student_dict:
            raise ExcelImportError(
                f"Student '{first_name} {last_name}' not found",
                "Load the students file before loading classrooms.",
            )

        teacher = teacher_dict[teacher_name]
        student = student_dict[student_key]

        if teacher.name in classrooms:
            classrooms[teacher.name].students.append(student)
        else:
            classrooms[teacher.name] = Classroom(teacher, [student])

    return list(classrooms.values())


def _sheet_to_clean_records(file: pd.ExcelFile, sheet_name: str) -> list[dict[str, Any]]:
    """Parse an Excel sheet into a list of clean dicts.

    Drops rows with all NaN values and converts each row to a dict.

    Args:
        file: Open ExcelFile object.
        sheet_name: Name of the sheet to read.

    Returns:
        List of dicts, one per non-empty row.

    Raises:
        ExcelImportError: If no sheet can be found.
    """
    if sheet_name in file.sheet_names:
        df = file.parse(sheet_name=sheet_name)
    elif len(file.sheet_names):
        df = file.parse(sheet_name=0)
    else:
        raise ExcelImportError("Could not find a sheet in the provided excel file")

    records: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        clean_row: dict[str, Any] = {str(k): v for k, v in row.dropna().to_dict().items()}
        if clean_row:
            records.append(clean_row)
    return records
